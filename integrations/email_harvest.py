"""Извлечение email из текста, PDF, Excel, CSV (для списка рассылки)."""

from __future__ import annotations

import io
import re

EMAIL_PATTERN = re.compile(
    r"[a-zA-Z0-9][a-zA-Z0-9._%+-]*@[a-zA-Z0-9][a-zA-Z0-9.-]*\.[a-zA-Z]{2,}",
    re.IGNORECASE,
)

_MAX_EXTRACT_BYTES = 20 * 1024 * 1024


def extract_emails_from_text(text: str) -> list[str]:
    if not text:
        return []
    seen: set[str] = set()
    out: list[str] = []
    for m in EMAIL_PATTERN.finditer(text):
        e = m.group(0).lower().rstrip(".,;:)\"'»")
        if e not in seen and "@" in e:
            seen.add(e)
            out.append(e)
    return out


def extract_emails_from_pdf_bytes(data: bytes) -> list[str]:
    import pdfplumber

    chunks: list[str] = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                chunks.append(t)
            for table in page.extract_tables() or []:
                for row in table:
                    if not row:
                        continue
                    for cell in row:
                        if cell:
                            chunks.append(str(cell))
    return extract_emails_from_text("\n".join(chunks))


def extract_emails_from_xlsx_bytes(data: bytes) -> list[str]:
    from openpyxl import load_workbook

    chunks: list[str] = []
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    v = cell.value
                    if v is not None and str(v).strip():
                        chunks.append(str(v))
    finally:
        wb.close()
    return extract_emails_from_text("\n".join(chunks))


def extract_emails_from_xls_bytes(data: bytes) -> list[str]:
    import xlrd

    chunks: list[str] = []
    book = xlrd.open_workbook(file_contents=data)
    for si in range(book.nsheets):
        sheet = book.sheet_by_index(si)
        for ri in range(sheet.nrows):
            for ci in range(sheet.ncols):
                v = sheet.cell_value(ri, ci)
                if v is not None and str(v).strip():
                    chunks.append(str(v))
    return extract_emails_from_text("\n".join(chunks))


def extract_emails_from_csv_bytes(data: bytes) -> list[str]:
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = data.decode(enc)
            return extract_emails_from_text(text)
        except UnicodeDecodeError:
            continue
    return extract_emails_from_text(data.decode("utf-8", errors="replace"))


def extract_emails_from_upload(filename: str, data: bytes) -> tuple[list[str], str]:
    if len(data) > _MAX_EXTRACT_BYTES:
        raise ValueError(f"Файл больше {_MAX_EXTRACT_BYTES // (1024 * 1024)} МБ")
    name = (filename or "").lower().strip()
    if name.endswith(".pdf"):
        return extract_emails_from_pdf_bytes(data), "pdf"
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return extract_emails_from_xlsx_bytes(data), "xlsx"
    if name.endswith(".xls"):
        return extract_emails_from_xls_bytes(data), "xls"
    if name.endswith(".csv") or name.endswith(".txt"):
        return extract_emails_from_csv_bytes(data), "csv"
    raise ValueError(
        "Поддерживаются: .pdf, .xlsx, .xls, .xlsm, .csv, .txt — или укажите ссылку на Google Таблицу"
    )
